"""
Magic: The Gathering - Final Fantasy Collection Tracker Generator
Generates:
1. MTG_Final_Fantasy_Collection_Tracker.xlsx (Multi-tab Excel / Google Sheets workbook)
2. MTG_Final_Fantasy_Collection.csv (Direct CSV import file with =IMAGE formulas)
3. MTG_Final_Fantasy_Collection.tsv (Tab-separated import file)
4. cards_data.js (the card database that index.html / app.js read)

Usage:
    python generate_ff_tracker.py             regenerate every output
    python generate_ff_tracker.py --only-js   regenerate cards_data.js only (used by CI)
    python generate_ff_tracker.py --dry-run   write *.new files, leave the real ones alone
"""

import argparse
import json
import os
from urllib.parse import quote
import sys
import time
import re
import csv
from typing import List, Dict, Any
import requests
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

# ---------------------------------------------------------------------------
# Configuration & Constants
# ---------------------------------------------------------------------------

# The nine sets that are wholly Final Fantasy.
FF_SET_CODES = ["fin", "fic", "fca", "afic", "afin", "rfin", "pfin", "tfin", "tfic"]

# Asking only for those nine sets misses Final Fantasy cards printed elsewhere:
# Pro Tour promos (PPRO), media inserts (PMEI), Spotlight Series (PSPL), Secret
# Lair (SLD) and several store/event promos all contain FF cards under their own
# set codes. Eighteen printings were missing because of this.
#
# "in:fin" means "every printing of any card that appears in FIN", so this finds
# those strays wherever they live. It also drags in ordinary Magic printings of
# the cards FCA reprints under Final Fantasy names, which is what
# is_final_fantasy_printing() below filters back out.
SCRYFALL_API_URL = (
    "https://api.scryfall.com/cards/search?"
    "q=" + quote("game:paper (" + " or ".join("in:" + code for code in FF_SET_CODES) + ")") +
    "&unique=prints&include_extras=true"
)

USER_AGENT = "MTG-Final-Fantasy-Collection-Tracker/1.0 (Contact: collector@example.com)"
OUTPUT_XLSX = "MTG_Final_Fantasy_Collection_Tracker.xlsx"
OUTPUT_CSV = "MTG_Final_Fantasy_Collection.csv"
OUTPUT_TSV = "MTG_Final_Fantasy_Collection.tsv"
OUTPUT_CARDS_JS = "cards_data.js"
OUTPUT_PRICE_HISTORY = "price_history.js"

# How many readings to keep per card. The refresh runs weekly, so this is a
# rolling year. Older readings fall off the front rather than growing the file
# without limit - at roughly 13 KB per reading, a year is about 700 KB.
HISTORY_LIMIT = 52

MAIN_SHEET_COLUMNS = [
    "Final Fantasy Card Name",
    "Original MTG Name",
    "Card Image",
    "Card Set",
    "Card Number",
    "Print Variant",
    "Available Variants",
    "Treatment / Frame",
    "Game Number",
    "Rarity",
    "Type Line",
    "Color Identity",
    "Owned",
    "Non-Foil Qty",
    "Foil Qty",
    "Special Foil Qty",
    "Total Quantity Owned",
    "Condition",
    "Price (USD)",
    "Price (Foil)",
    "Storage Location",
]

VALIDATION_OPTIONS = {
    "Card Set": ["FIN", "FIC", "FCA", "AFIC", "AFIN", "RFIN", "PFIN", "TFIN", "TFIC"],
    "Print Variant": [
        "Basic/Non-foil",
        "Traditional Foil",
        "Surge Foil",
        "Wave Foil",
        "Foil Etched",
        "Promo",
        "Serialized",
    ],
    # Wizards' own treatment names, taken from the Card Image Gallery's CMS.
    # See determine_treatment() for where the list comes from.
    "Treatment / Frame": [
        "Showcase",
        "Borderless",
        "Extended Art",
        "Full Art",
        "Default",
        "Art Card",
    ],
    "Game Number": [
        "FF1",
        "FF2",
        "FF3",
        "FF4",
        "FF5",
        "FF6",
        "FF7",
        "FF8",
        "FF9",
        "FF10",
        "FF11",
        "FF12",
        "FF13",
        "FF14",
        "FF15",
        "FF16",
        "Spin-Off / Multi-Game",
        "Unknown",
    ],
    "Owned": ["Yes", "No"],
    "Condition": [
        "Near Mint (NM)",
        "Lightly Played (LP)",
        "Moderately Played (MP)",
        "Heavily Played (HP)",
        "Damaged (DMG)",
    ],
}

ROMAN_TO_GAME = {
    "ffi": "FF1", "ffii": "FF2", "ffiii": "FF3", "ffiv": "FF4",
    "ffv": "FF5", "ffvi": "FF6", "ffvii": "FF7", "ffviii": "FF8",
    "ffix": "FF9", "ffx": "FF10", "ffxi": "FF11", "ffxii": "FF12",
    "ffxiii": "FF13", "ffxiv": "FF14", "ffxv": "FF15", "ffxvi": "FF16"
}

LORE_KEYWORDS = {
    "FF1": ["garland", "chaos", "warrior of light", "princess sarah", "matoya", "bikke", "astos", "lukahn", "pravoka", "onrac", "lufenia", "cornelia"],
    "FF2": ["firion", "maria", "guy", "leon", "minwu", "gordon", "josef", "leila", "ricard", "emperor mateus", "hilda", "dreadnought", "paramekia", "pandemonium", "wild rose"],
    "FF3": ["luneth", "arc", "refia", "ingus", "cloud of darkness", "xande", "doga", "unei", "desch", "aria", "hein", "crystal tower", "floating continent"],
    "FF4": ["cecil", "kain", "rosa", "rydia", "edge", "golbez", "zeromus", "yang", "palom", "porom", "edward", "tellah", "fusoya", "cid pollan", "rubicante", "cagnazzo", "barbariccia", "scarmiglione", "lunar whale", "baron"],
    "FF5": ["bartz", "lenna", "galuf", "faris", "krile", "exdeath", "gilgamesh", "enuo", "shinryu", "omega", "ghido", "tycoon", "void", "syldra"],
    "FF6": ["terra", "locke", "celes", "edgar", "sabin", "shadow", "cyan", "gau", "setzer", "strago", "relm", "mog", "umaro", "gogo", "kefka", "gestahl", "general leo", "opera house", "narshe", "figaro", "vector", "esper"],
    "FF7": ["cloud", "tifa", "aerith", "barret", "sephiroth", "red xiii", "yuffie", "vincent", "cait sith", "cid highwind", "zack", "midgar", "shinra", "jenova", "materia", "chocobo", "buster sword", "gold saucer", "sector 7", "nibelheim", "hojo", "rufus", "avalanche", "cosmo canyon"],
    "FF8": ["squall", "rinoa", "seifer", "zell", "irvine", "quistis", "selphie", "laguna", "kiros", "ward", "edea", "ultimecia", "balamb", "seed", "galbadia", "esthar", "sorceress", "gunblade", "lionheart", "ragnarok", "triple triad"],
    "FF9": ["zidane", "garnet", "dagger", "vivi", "steiner", "freya", "quina", "eiko", "amarant", "kuja", "brahne", "beatrix", "alexandria", "lindblum", "burmecia", "mist continent", "prima vista", "trance", "eidolons"],
    "FF10": ["tidus", "yuna", "auron", "wakka", "lulu", "kimahri", "rikku", "seymour", "sin", "jecht", "spira", "zanarkand", "besaid", "kilika", "luca", "blitzball", "bevelle", "aeon", "calm lands", "macalania", "gagazet", "yu yevon"],
    "FF11": ["shantotto", "ayame", "curilla", "zeid", "prishe", "aphmau", "lilisette", "arciela", "shadow lord", "vana'diel", "san d'oria", "bastok", "windurst", "jeuno", "aht urhgan", "absolute virtue", "tarutaru", "elvaan", "mithra", "galka", "hume"],
    "FF12": ["vaan", "penelo", "balthier", "fran", "basch", "ashe", "vayne", "gabranth", "cidolfus", "dalmasca", "rabanastre", "archadia", "strahl", "yiazmat", "occuria", "nethicite", "judge magister", "skypirate"],
    "FF13": ["lightning", "snow", "vanille", "fang", "hope", "sazh", "serah", "noel", "caius", "barthandelus", "bhunivelze", "cocoon", "pulse", "fal'cie", "l'cie", "eidolons", "bodhum", "eden"],
    "FF14": ["alphinaud", "alisaie", "yshtola", "y'shtola", "thancred", "urianger", "estinien", "graha", "g'raha", "scions", "hydaelyn", "zodiark", "emet-selch", "elidibus", "lahabrea", "zenos", "hades", "eorzea", "scion", "warrior of light", "venat", "meteion", "limsa", "gridania", "uldah", "ishgard", "kugane", "scion of the seventh dawn", "a realm reborn"],
    "FF15": ["noctis", "gladiolus", "ignis", "prompto", "lunafreya", "ardyn", "cor", "regalia", "lucis", "insomnia", "niflheim", "astral", "armiger", "cindy", "chocobo post", "king of lucis", "ring of the lucii"],
    "FF16": ["clive", "joshua", "jill", "dion", "barnabas", "benedikta", "hugo", "cidolfus", "torgal", "ifrit", "phoenix", "garuda", "titan", "ramuh", "bahamut", "odin", "ultima", "valisthea", "rosaria", "sanbreque", "dhalmekia", "walood", "dominant", "eikon", "bearer", "mothercrystal"]
}


# ---------------------------------------------------------------------------
# API & Data Extraction Functions
# ---------------------------------------------------------------------------

def fetch_scryfall_cards(query_url: str = SCRYFALL_API_URL) -> List[Dict[str, Any]]:
    headers = {
        "User-Agent": USER_AGENT,
        "Accept": "application/json",
    }
    
    cards: List[Dict[str, Any]] = []
    current_url = query_url
    page_num = 1
    
    print("=" * 65)
    print("Fetching Final Fantasy cards from Scryfall API...")
    print("=" * 65)
    
    while current_url:
        print(f"Fetching page {page_num}...")
        # The full query runs to about sixty pages, and Scryfall starts
        # answering 429 partway through even at its documented pacing. Back off
        # and retry rather than losing the whole run.
        data = None
        for attempt in range(6):
            try:
                response = requests.get(current_url, headers=headers, timeout=30)
                response.raise_for_status()
                data = response.json()
                break
            except requests.exceptions.RequestException as e:
                wait = 2 ** attempt
                print(f"  ! page {page_num} failed ({e}); retrying in {wait}s")
                time.sleep(wait)
        if data is None:
            print(f"Error fetching page {page_num} from Scryfall: gave up after 6 attempts")
            raise SystemExit(1)

        page_cards = data.get("data", [])
        cards.extend(page_cards)
        print(f"  -> Page {page_num} parsed: {len(page_cards)} cards (Running Total: {len(cards)})")
        
        has_more = data.get("has_more", False)
        current_url = data.get("next_page") if has_more else None
        page_num += 1
        
        if current_url:
            time.sleep(0.25)
            
    print(f"\nAll pages fetched successfully. Total printings retrieved: {len(cards)}")

    ff_cards = [card for card in cards if is_final_fantasy_printing(card)]
    dropped = len(cards) - len(ff_cards)
    print(f"Kept {len(ff_cards)} Final Fantasy printings ({dropped} unrelated printings filtered out).")
    return ff_cards


def is_final_fantasy_printing(card: Dict[str, Any]) -> bool:
    """Keep a printing only if it is genuinely a Final Fantasy card.

    Two ways to qualify:
      1. it lives in one of the nine Final Fantasy sets, or
      2. it carries a Final Fantasy game tag (ffvii and friends), which is how
         the strays in PPRO / PMEI / PSPL / SLD identify themselves.

    This is what removes the ordinary Magic printings that "in:fca" drags in.
    FCA reprints existing cards under Final Fantasy names, so asking for every
    printing of them also returns the originals from their real sets.
    """
    if (card.get("set") or "").lower() in FF_SET_CODES:
        return True
    tags = {str(tag).lower() for tag in (card.get("promo_types") or [])}
    return bool(tags & set(ROMAN_TO_GAME.keys()))


def determine_ff_game(card: Dict[str, Any], name_lookup: Dict[str, str]) -> str:
    promo_types = card.get("promo_types") or []
    for promo in promo_types:
        promo_lower = promo.lower()
        if promo_lower in ROMAN_TO_GAME:
            return ROMAN_TO_GAME[promo_lower]
    
    full_name = card.get("name", "").strip()
    if full_name in name_lookup:
        return name_lookup[full_name]
    
    if " // " in full_name:
        for part in full_name.split(" // "):
            part_clean = part.strip()
            if part_clean in name_lookup:
                return name_lookup[part_clean]
    
    search_corpus = " ".join([
        full_name,
        card.get("flavor_name") or "",
        card.get("flavor_text") or "",
        card.get("oracle_text") or "",
        card.get("type_line") or "",
    ]).lower()
    
    for game, keywords in LORE_KEYWORDS.items():
        for kw in keywords:
            if re.search(r"\b" + re.escape(kw) + r"\b", search_corpus):
                return game
                
    layout = card.get("layout", "")
    type_line = (card.get("type_line") or "").lower()
    if layout in ["token", "double_faced_token", "emblem"] or "token" in type_line:
        return "Spin-Off / Multi-Game"
        
    return "Unknown"


def determine_print_variant(card: Dict[str, Any]) -> str:
    promo_types = card.get("promo_types") or []
    finishes = card.get("finishes") or []
    set_code = (card.get("set") or "").lower()
    
    if "serialized" in promo_types:
        return "Serialized"
    if "surgefoil" in promo_types:
        return "Surge Foil"
    if "chocobotrackfoil" in promo_types or "wavefoil" in promo_types:
        return "Wave Foil"
    if "etched" in finishes or "foiletched" in promo_types:
        return "Foil Etched"
    # card["promo"] catches the ones printed outside the Final Fantasy sets -
    # Pro Tour, media inserts, Spotlight Series - which carry none of the tags
    # listed below.
    if card.get("promo") or set_code in ["pfin", "rfin"] or any(
        p in promo_types for p in ["prerelease", "buyabox", "bundle", "datestamped", "starterdeck"]
    ):
        return "Promo"
    if finishes == ["foil"]:
        return "Traditional Foil"
    return "Basic/Non-foil"


# Wizards' official treatment vocabulary, read from the Card Image Gallery's
# CMS (Contentful space s5n2t79q9icq, content type "magicCard", field
# "treatments" -> tag entries titled "Magic Card Treatment | ...").
#
# The gallery at
#   https://magic.wizards.com/en/products/final-fantasy/card-image-gallery
# renders client-side, so the names are not in the served HTML; they come from
# that API. Twelve exist for Final Fantasy, and they split into two independent
# axes:
#
#   frame   Art Card, Showcase, Borderless, Extended Art, Full Art, Default
#   finish  Traditional Foil, Surge Foil, Chocobo Track Foil, Neon Ink, Serialized
#
# plus "Scene Card", a category that rides along with Borderless.
#
# Scryfall does not carry these names - `is:showcase` returns nothing for every
# Final Fantasy set - so the frame is derived below. The derivation was checked
# against all 1,258 cards that join to Wizards' data by set and collector
# number and reproduces their label for every one of them. Order matters:
# Borderless outranks Full Art (166 cards are both, and Wizards calls them
# Borderless), and tokens are Full Art even though Scryfall leaves full_art
# false on them.
TREATMENT_FRAMES = ["Showcase", "Borderless", "Extended Art", "Full Art", "Default", "Art Card"]

# The Scene Box holds 24 scenes, and each one exists twice: as a playable
# borderless card in the Commander set (a contiguous block, FIC 442-465) and as
# an art card in its own set, AFIC ("Final Fantasy Scene Box"). The two line up
# one for one by name.
#
# Wizards tags only the playable half "Scene Card" and calls the art half plain
# "Art Card" - but their own names for those read "<card> Scene Art Card N/24",
# so both halves are tagged here. That keeps the Scene Box findable as one thing
# while leaving the Art Card count alone.
SCENE_CARD_SET = "fic"
SCENE_CARD_RANGE = (442, 465)
SCENE_ART_SET = "afic"

# The Art Series proper. Kept separate from the Scene Box art cards: they come
# out of different products and sort as different things.
ART_SERIES_SET = "afin"


def determine_treatment(card: Dict[str, Any]) -> str:
    """The card's official Wizards frame treatment."""
    set_code = (card.get("set") or "").lower()
    layout = card.get("layout") or ""
    border = card.get("border_color") or ""
    frame_effects = card.get("frame_effects") or []
    type_line = card.get("type_line") or ""

    if layout == "art_series" or set_code in ["afin", "afic"]:
        return "Art Card"
    # The Through the Ages bonus sheet is Showcase in its entirety, and is the
    # only place in Final Fantasy the label is used outside a few promos.
    if set_code == "fca":
        return "Showcase"
    if border == "borderless":
        return "Borderless"
    if "extendedart" in frame_effects:
        return "Extended Art"
    if (card.get("full_art")
            or layout in ("token", "double_faced_token", "emblem")
            or set_code in ("tfin", "tfic")
            or "Token" in type_line):
        return "Full Art"
    return "Default"


def determine_treatments(card: Dict[str, Any]) -> List[str]:
    """Every official treatment that applies to this printing.

    Frame plus Scene Card plus the finishes. The finishes come from Scryfall
    rather than from Wizards: Wizards' own records omit them for the whole
    Commander set and for every art card, whereas Scryfall's promo_types and
    finishes are complete.
    """
    out = [determine_treatment(card)]

    set_code = (card.get("set") or "").lower()
    digits = re.sub(r"\D", "", str(card.get("collector_number") or ""))
    if set_code == SCENE_ART_SET:
        out.append("Scene Card")
    elif set_code == SCENE_CARD_SET and digits:
        if SCENE_CARD_RANGE[0] <= int(digits) <= SCENE_CARD_RANGE[1]:
            out.append("Scene Card")

    promo_types = set(card.get("promo_types") or [])
    finishes = set(card.get("finishes") or [])
    special = False
    if "surgefoil" in promo_types:
        out.append("Surge Foil"); special = True
    if "chocobotrackfoil" in promo_types:
        out.append("Chocobo Track Foil"); special = True
    if "neonink" in promo_types:
        out.append("Neon Ink"); special = True
    if "serialized" in promo_types or "serialized" in finishes:
        out.append("Serialized"); special = True
    if "foil" in finishes and not special:
        out.append("Traditional Foil")

    return out


def build_name_to_game(raw_cards: List[Dict[str, Any]]) -> Dict[str, str]:
    """Map card name -> FF game, seeded from the printings whose promo_types say so.

    Reprints in FCA carry no promo type of their own, so they are matched by name
    against a printing that does.
    """
    name_to_game: Dict[str, str] = {}
    for card in raw_cards:
        pts = card.get("promo_types") or []
        for promo in pts:
            promo_lower = promo.lower()
            if promo_lower in ROMAN_TO_GAME:
                game = ROMAN_TO_GAME[promo_lower]
                full_name = card.get("name", "").strip()
                name_to_game[full_name] = game
                if " // " in full_name:
                    for part in full_name.split(" // "):
                        name_to_game[part.strip()] = game
    return name_to_game


def parse_card_data(raw_cards: List[Dict[str, Any]]) -> pd.DataFrame:
    print("\nProcessing and structuring card metadata...")

    name_to_game = build_name_to_game(raw_cards)

    parsed_rows = []
    for card in raw_cards:
        mtg_name = card.get("name", "Unknown Card")
        flavor_name = card.get("flavor_name")
        
        face_flavor_names = []
        if "card_faces" in card:
            for face in card["card_faces"]:
                if face.get("flavor_name"):
                    face_flavor_names.append(face["flavor_name"])
                    
        if flavor_name:
            ff_name = flavor_name
        elif face_flavor_names:
            ff_name = " // ".join(face_flavor_names)
        else:
            ff_name = mtg_name
        
        image_url = ""
        if "image_uris" in card and card["image_uris"]:
            image_url = card["image_uris"].get("normal") or card["image_uris"].get("small") or ""
        elif "card_faces" in card and card["card_faces"] and len(card["card_faces"]) > 0:
            first_face = card["card_faces"][0]
            if "image_uris" in first_face and first_face["image_uris"]:
                image_url = first_face["image_uris"].get("normal") or first_face["image_uris"].get("small") or ""
        
        image_formula = f'=IMAGE("{image_url}", 1)' if image_url else ""
        
        set_code = (card.get("set") or "").upper()
        raw_num = str(card.get("collector_number", ""))
        collector_number = raw_num.zfill(4) if raw_num.isdigit() else raw_num
        
        variant = determine_print_variant(card)
        treatment = determine_treatment(card)
        game_number = determine_ff_game(card, name_to_game)
        
        rarity = (card.get("rarity") or "Unknown").capitalize()
        type_line = card.get("type_line", "")
        
        color_id_list = card.get("color_identity", [])
        color_identity = "".join(color_id_list) if color_id_list else "Colorless"
        
        prices = card.get("prices") or {}
        price_usd_str = prices.get("usd")
        price_foil_str = prices.get("usd_foil")
        
        price_usd = float(price_usd_str) if price_usd_str else None
        price_foil = float(price_foil_str) if price_foil_str else None
        
        finishes = card.get("finishes") or []
        promo_types = card.get("promo_types") or []
        set_lower = set_code.lower()
        
        if "serialized" in promo_types:
            avail_vars = "Serialized"
        elif "surgefoil" in promo_types:
            avail_vars = "Surge Foil"
        elif "chocobotrackfoil" in promo_types or "wavefoil" in promo_types:
            avail_vars = "Wave Foil"
        elif "etched" in finishes or "foiletched" in promo_types:
            avail_vars = "Foil Etched"
        elif set_lower in ["pfin", "rfin"] or any(p in promo_types for p in ["prerelease", "buyabox", "bundle", "datestamped"]):
            v_list = []
            if "nonfoil" in finishes:
                v_list.append("Non-Foil")
            if "foil" in finishes:
                v_list.append("Traditional Foil")
            v_list.append("Promo")
            avail_vars = ", ".join(v_list)
        else:
            v_list = []
            if "nonfoil" in finishes:
                v_list.append("Non-Foil")
            if "foil" in finishes:
                v_list.append("Traditional Foil")
            avail_vars = ", ".join(v_list) if v_list else "Non-Foil"

        parsed_rows.append({
            "Final Fantasy Card Name": ff_name,
            "Original MTG Name": mtg_name if ff_name != mtg_name else "",
            "Card Image": image_formula,
            "Card Set": set_code,
            "Card Number": collector_number,
            "Print Variant": variant,
            "Available Variants": avail_vars,
            "Treatment / Frame": treatment,
            "Game Number": game_number,
            "Rarity": rarity,
            "Type Line": type_line,
            "Color Identity": color_identity,
            "Owned": "No",
            "Non-Foil Qty": 0,
            "Foil Qty": 0,
            "Special Foil Qty": 0,
            "Total Quantity Owned": 0,
            "Condition": "Near Mint (NM)",
            "Price (USD)": price_usd,
            "Price (Foil)": price_foil,
            "Storage Location": "",
        })
        
    df = pd.DataFrame(parsed_rows)
    print(f"Data structuring complete: {len(df)} cards processed.")
    return df


# ---------------------------------------------------------------------------
# Web tracker data export (cards_data.js)
#
# index.html loads cards_data.js as a plain script that defines a global
# CARDS_DATA array. Keeping it as JS rather than JSON means the site needs no
# fetch() and therefore works when opened straight off the filesystem.
# ---------------------------------------------------------------------------

def _repoint_self_reference(text: str, mtg_name: str, ff_name: str) -> str:
    """Point a reprint's rules text at the character actually on the card.

    A fallback for the handful of cards Scryfall has no printed_text for. Card
    text refers to the card by its short name - everything before the first
    comma - so "Inalla, Archmage Ritualist" is "Inalla" in its own text, and
    the Final Fantasy printing of it reads "Kuja". Verified against the card
    face for FCA 0052, which is the only card this currently fires on.

    Deliberately narrow: whole words only, and only when the two short names
    actually differ. Anything more ambitious would start rewriting rules text
    that mentions another card by name.
    """
    if not text or not mtg_name or not ff_name or mtg_name == ff_name:
        return text
    mtg_short = re.split(r"[,(]| // ", mtg_name)[0].strip()
    ff_short = re.split(r"[,(]| // ", ff_name)[0].strip()
    if not mtg_short or mtg_short == ff_short:
        return text
    updated = re.sub(r"\b%s\b" % re.escape(mtg_short), ff_short, text)
    if updated != text:
        print(f"  Rules text repointed: {mtg_short} -> {ff_short} ({ff_name})")
    return updated


def _face_values(card: Dict[str, Any], key: str) -> List[str]:
    """Collect a field from each card face, skipping blanks."""
    values = []
    for face in card.get("card_faces") or []:
        value = face.get(key)
        if value:
            values.append(value)
    return values


def _first_image_uris(card: Dict[str, Any]) -> Dict[str, str]:
    """Front-face images. Double-faced cards carry them per face, not top level."""
    if card.get("image_uris"):
        return card["image_uris"]
    faces = card.get("card_faces") or []
    if faces and faces[0].get("image_uris"):
        return faces[0]["image_uris"]
    return {}


def _back_image(card: Dict[str, Any]) -> str:
    faces = card.get("card_faces") or []
    if len(faces) > 1 and faces[1].get("image_uris"):
        return faces[1]["image_uris"].get("normal") or ""
    return ""


def _price(prices: Dict[str, Any], key: str):
    raw = prices.get(key)
    try:
        return float(raw) if raw else None
    except (TypeError, ValueError):
        return None


def _available_variants(card: Dict[str, Any]) -> List[str]:
    """Which finishes THIS printing exists in, as a list.

    Mirrors the 'Available Variants' string built in parse_card_data, so the
    spreadsheet and the web tracker never disagree.
    """
    promo_types = card.get("promo_types") or []
    finishes = card.get("finishes") or []
    set_lower = (card.get("set") or "").lower()

    # Art cards are never printed in foil - Wizards gives them a gloss varnish
    # instead. The second version is the artist's gold-stamped signature, which
    # Scryfall records as the "foil" finish for want of anywhere better. So the
    # two versions are Basic and Signed, not Non-Foil and Traditional Foil.
    #
    # Both names map back to the nonfoil/foil storage keys in app.js, so a
    # collection recorded before this change keeps its counts.
    if (card.get("layout") or "") == "art_series" or set_lower in (ART_SERIES_SET, SCENE_ART_SET):
        variants = []
        if "nonfoil" in finishes:
            variants.append("Basic")
        if "foil" in finishes:
            variants.append("Signed")
        return variants or ["Basic"]

    if "serialized" in promo_types:
        return ["Serialized"]
    if "surgefoil" in promo_types:
        return ["Surge Foil"]
    if "chocobotrackfoil" in promo_types or "wavefoil" in promo_types:
        return ["Wave Foil"]
    if "etched" in finishes or "foiletched" in promo_types:
        return ["Foil Etched"]

    variants: List[str] = []
    if "nonfoil" in finishes:
        variants.append("Non-Foil")
    if "foil" in finishes:
        variants.append("Traditional Foil")

    is_promo = card.get("promo") or set_lower in ["pfin", "rfin"] or any(
        p in promo_types for p in ["prerelease", "buyabox", "bundle", "datestamped"]
    )
    if is_promo:
        variants.append("Promo")

    return variants or ["Non-Foil"]


def build_card_records(raw_cards: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Turn raw Scryfall cards into the record shape app.js expects."""
    print("\nBuilding web tracker card records...")

    name_to_game = build_name_to_game(raw_cards)
    records: List[Dict[str, Any]] = []

    for card in raw_cards:
        mtg_name = card.get("name", "Unknown Card")

        flavor_name = card.get("flavor_name")
        face_flavor_names = _face_values(card, "flavor_name")
        if flavor_name:
            ff_name = flavor_name
        elif face_flavor_names:
            ff_name = " // ".join(face_flavor_names)
        else:
            ff_name = mtg_name

        raw_num = str(card.get("collector_number", ""))
        collector_number = raw_num.zfill(4) if raw_num.isdigit() else raw_num

        color_id_list = card.get("color_identity") or []
        color_identity = "".join(color_id_list) if color_id_list else "Colorless"

        images = _first_image_uris(card)
        prices = card.get("prices") or {}

        # Double-faced cards keep type line / mana cost / rules text per face.
        # The old data file left these blank for every DFC; join the faces so
        # the modal and the search actually have something to work with.
        mana_cost = card.get("mana_cost") or " // ".join(_face_values(card, "mana_cost"))

        # printed_* before the Oracle fields, in every language.
        #
        # Oracle text is the rules text filed under a card's MAGIC name, so a
        # Final Fantasy card that reprints one self-references the wrong
        # character: "Adeline's power is equal to..." where the card in your
        # hand reads "Hero of Light's power is equal to...". The printed_*
        # fields are what is actually on the card, which is the whole point.
        #
        # This applies to non-English printings too: a Japanese card shows its
        # Japanese rules text and type line, matching the Japanese name it
        # already displays. Most cards have no printed_* fields at all and fall
        # through to Oracle unchanged.
        type_line = (card.get("printed_type_line")
                     or card.get("type_line")
                     or " // ".join(_face_values(card, "printed_type_line")
                                    or _face_values(card, "type_line")))
        printed = (card.get("printed_text")
                   or "\n//\n".join(_face_values(card, "printed_text")))
        oracle_text = (printed
                       or card.get("oracle_text")
                       or "\n//\n".join(_face_values(card, "oracle_text")))
        if not printed:
            oracle_text = _repoint_self_reference(oracle_text, mtg_name, ff_name)
        flavor_text = card.get("flavor_text") or ""
        if not flavor_text:
            face_flavor = _face_values(card, "flavor_text")
            flavor_text = face_flavor[0] if face_flavor else ""
        artist = card.get("artist") or " & ".join(dict.fromkeys(_face_values(card, "artist")))

        records.append({
            "id": card.get("id"),
            "name": ff_name,
            "ff_name": ff_name,
            "mtg_name": mtg_name,
            "is_reprint": ff_name != mtg_name,
            "set": (card.get("set") or "").upper(),
            "set_name": card.get("set_name") or "",
            "collector_number": collector_number,
            "variant": determine_print_variant(card),
            "avail_variants": _available_variants(card),
            "treatment": determine_treatment(card),
            "treatments": determine_treatments(card),
            "game": determine_ff_game(card, name_to_game),
            "rarity": (card.get("rarity") or "Unknown").capitalize(),
            "type_line": type_line,
            "mana_cost": mana_cost,
            "oracle_text": oracle_text,
            "flavor_text": flavor_text,
            "artist": artist,
            "color_identity": color_identity,
            "finishes": card.get("finishes") or [],
            "promo_types": card.get("promo_types") or [],
            "image_small": images.get("small", ""),
            "image_normal": images.get("normal", ""),
            "image_large": images.get("large", ""),
            "back_image": _back_image(card),
            "price_usd": _price(prices, "usd"),
            "price_foil": _price(prices, "usd_foil"),
            "price_etched": _price(prices, "usd_etched"),
            # Scryfall has NO dollar price for the Art Series, the Scene Box or a
            # handful of promos - only euros. One of those is a Cloud promo worth
            # several hundred. Carrying the euro price means the site can show a
            # real number for them instead of "--".
            "price_eur": _price(prices, "eur"),
            "price_eur_foil": _price(prices, "eur_foil"),
            "scryfall_uri": card.get("scryfall_uri", ""),
        })

    # Deterministic order keeps the monthly price-refresh diff readable: a price
    # change should be a one-line diff, not a reshuffled file.
    def sort_key(record):
        digits = re.sub(r"\D", "", record["collector_number"])
        return (record["set"], int(digits) if digits else 0, record["collector_number"])

    records.sort(key=sort_key)
    print(f"Card records built: {len(records)}")
    return records


def export_cards_js(records: List[Dict[str, Any]], output_path: str = OUTPUT_CARDS_JS) -> None:
    print(f"\nWriting web tracker card database: '{output_path}'...")
    payload = json.dumps(records, indent=2, ensure_ascii=False)
    banner = (
        "/**\n"
        " * Card database for the Final Fantasy MTG Collection Tracker.\n"
        " *\n"
        " * GENERATED FILE - do not edit by hand.\n"
        " * Regenerate with:  python generate_ff_tracker.py --only-js\n"
        " *\n"
        " * Source: Scryfall (https://scryfall.com). Card images are hotlinked from\n"
        " * Scryfall's CDN rather than rehosted.\n"
        " */\n"
    )
    with open(output_path, "w", encoding="utf-8", newline="\n") as handle:
        handle.write(banner)
        handle.write("const CARDS_DATA = ")
        handle.write(payload)
        handle.write(";\n")

    size_mb = os.path.getsize(output_path) / (1024 * 1024)
    print(f"Card database written: {len(records)} cards, {size_mb:.2f} MB.")


def _read_price_history(path: str) -> Dict[str, Any]:
    """Load price_history.js, or start a fresh one.

    The file is a plain `const PRICE_HISTORY = {...};` script so it can be
    loaded straight from the page - and from a file:// copy - with no fetch and
    no server, exactly like cards_data.js.
    """
    if not os.path.exists(path):
        return {"dates": [], "cards": {}}
    with open(path, encoding="utf-8") as handle:
        text = handle.read()
    start = text.find("{", text.find("PRICE_HISTORY"))
    if start == -1:
        print("  ! price history file is unreadable; starting a new one")
        return {"dates": [], "cards": {}}
    try:
        data = json.loads(text[start:].rstrip().rstrip(";"))
    except ValueError as exc:
        print(f"  ! price history file is unreadable ({exc}); starting a new one")
        return {"dates": [], "cards": {}}
    data.setdefault("dates", [])
    data.setdefault("cards", {})
    return data


def update_price_history(records: List[Dict[str, Any]],
                         path: str = OUTPUT_PRICE_HISTORY,
                         today: str = None) -> Dict[str, Any]:
    """Record today's prices and rewrite price_history.js.

    Shape:

        { "dates": ["2026-08-24", ...],
          "cards": { "<scryfall id>": { "usd": [...], "foil": [...] } } }

    One shared date list, and one reading per date per card, so a card id is
    stored once rather than once per snapshot. A reading with no price is null.

    Re-running on a day already recorded OVERWRITES that day rather than adding
    a second column, so a manual run cannot double up a date.

    Only the last HISTORY_LIMIT dates are kept. Weekly, that is a rolling year.
    """
    if today is None:
        today = time.strftime("%Y-%m-%d", time.gmtime())

    history = _read_price_history(path)
    dates: List[str] = history["dates"]
    cards: Dict[str, Any] = history["cards"]

    if today in dates:
        index = dates.index(today)
        rewriting = True
    else:
        dates.append(today)
        index = len(dates) - 1
        rewriting = False

    width = len(dates)

    def price_of(record, key):
        value = record.get(key)
        return round(float(value), 2) if isinstance(value, (int, float)) else None

    seen = set()
    for record in records:
        usd = price_of(record, "price_usd")
        foil = price_of(record, "price_foil")
        etched = price_of(record, "price_etched")
        # The euro price is only recorded where there is no dollar one for the
        # same finish. Keeping both everywhere would double the file to say the
        # same thing twice - and the site shows the euro line only when the
        # dollar line is missing, so this is exactly what gets displayed.
        eur = price_of(record, "price_eur") if usd is None else None
        eur_foil = price_of(record, "price_eur_foil") if foil is None else None

        readings = (("usd", usd), ("foil", foil), ("etched", etched),
                    ("eur", eur), ("eurFoil", eur_foil))

        if all(value is None for _, value in readings) and record["id"] not in cards:
            # No price in any currency, and never has had one. Tokens are the
            # bulk of these. A row would be a column of nulls; if a price ever
            # appears, the row is created then.
            continue

        seen.add(record["id"])
        series = cards.setdefault(record["id"], {})
        for name, value in readings:
            # Only carry a column once it has something in it, so a card priced
            # in euros does not also store three empty dollar columns.
            if name not in series and value is None:
                continue
            column = series.setdefault(name, [])
            while len(column) < width:
                column.append(None)
            column[index] = value

    # A card that has dropped out of the data keeps its past readings, but gets
    # a null for today rather than a stale repeat of last week's price.
    for card_id, series in cards.items():
        if card_id in seen:
            continue
        for name in list(series.keys()):
            column = series[name]
            while len(column) < width:
                column.append(None)

    # Trim to the rolling window.
    if width > HISTORY_LIMIT:
        cut = width - HISTORY_LIMIT
        history["dates"] = dates[cut:]
        for series in cards.values():
            for name in list(series.keys()):
                series[name] = series[name][cut:]

    # Drop cards whose whole remaining window is empty.
    # Drop empty columns first, then any card left with nothing at all.
    for series in cards.values():
        for name in list(series.keys()):
            if all(v is None for v in series[name]):
                del series[name]
    history["cards"] = {
        card_id: series for card_id, series in cards.items() if series
    }

    banner = (
        "/**\n"
        " * Price history for the Final Fantasy MTG Collection Tracker.\n"
        " *\n"
        " * GENERATED FILE - do not edit by hand. Appended to on every refresh;\n"
        " * past readings are never rewritten, so this is the one generated file\n"
        " * whose history would be lost if it were deleted.\n"
        " *\n"
        " * dates: the day of each reading, oldest first.\n"
        " * cards: scryfall id -> { usd: [...], foil: [...] }, one entry per date,\n"
        f" *        null where there was no price. Keeps the last {HISTORY_LIMIT} readings.\n"
        " */\n"
    )
    with open(path, "w", encoding="utf-8", newline="\n") as handle:
        handle.write(banner)
        handle.write("const PRICE_HISTORY = ")
        handle.write(json.dumps(history, separators=(",", ":"), ensure_ascii=False))
        handle.write(";\n")

    size_kb = os.path.getsize(path) / 1024
    verb = "rewrote" if rewriting else "added"
    print(f"Price history: {verb} {today}; "
          f"{len(history['dates'])} readings, {len(history['cards'])} cards, {size_kb:.0f} KB.")
    return history


# ---------------------------------------------------------------------------
# Excel Workbook & CSV/TSV Exporters
# ---------------------------------------------------------------------------

def export_csv_and_tsv(df: pd.DataFrame, csv_path: str = OUTPUT_CSV, tsv_path: str = OUTPUT_TSV) -> None:
    print(f"\nExporting direct Google Sheets import files: '{csv_path}' & '{tsv_path}'...")
    df.to_csv(csv_path, index=False, quoting=csv.QUOTE_NONNUMERIC, encoding="utf-8-sig")
    df.to_csv(tsv_path, index=False, sep="\t", quoting=csv.QUOTE_NONNUMERIC, encoding="utf-8-sig")
    print("CSV and TSV export complete.")


def build_excel_tracker(df: pd.DataFrame, output_path: str = OUTPUT_XLSX) -> None:
    print("\nBuilding Excel / Google Sheets workbook (.xlsx) with openpyxl...")
    wb = openpyxl.Workbook()
    
    ws_main = wb.active
    ws_main.title = "FF MTG Collection"
    ws_dash = wb.create_sheet(title="Collection Dashboard")
    ws_val = wb.create_sheet(title="_Validation")
    ws_val.sheet_state = "hidden"
    
    ws_main.views.sheetView[0].showGridLines = True
    ws_dash.views.sheetView[0].showGridLines = True
    
    val_col_ranges = {}
    for col_idx, (key, options) in enumerate(VALIDATION_OPTIONS.items(), start=1):
        col_letter = get_column_letter(col_idx)
        ws_val.cell(row=1, column=col_idx, value=key)
        for row_idx, opt in enumerate(options, start=2):
            ws_val.cell(row=row_idx, column=col_idx, value=opt)
        val_col_ranges[key] = f"=_Validation!${col_letter}$2:${col_letter}${len(options) + 1}"
        
    cols_to_write = [col for col in MAIN_SHEET_COLUMNS if col in df.columns]
    
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border_side = Side(style="thin", color="CBD5E1")
    cell_border = Border(
        left=thin_border_side, right=thin_border_side,
        top=thin_border_side, bottom=thin_border_side
    )
    
    ws_main.row_dimensions[1].height = 32
    for col_idx, col_name in enumerate(cols_to_write, start=1):
        cell = ws_main.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = cell_border
        
    data_font = Font(name="Segoe UI", size=10)
    data_bold_font = Font(name="Segoe UI", size=10, bold=True)
    
    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    total_rows = len(df)
    print(f"Writing {total_rows} card rows to 'FF MTG Collection'...")
    
    for r_idx, row in df.iterrows():
        excel_row = r_idx + 2
        ws_main.row_dimensions[excel_row].height = 80
        
        for c_idx, col_name in enumerate(cols_to_write, start=1):
            val = row[col_name]
            cell = ws_main.cell(row=excel_row, column=c_idx, value=val)
            cell.font = data_font
            cell.border = cell_border
            
            if col_name in ["Final Fantasy Card Name", "Original MTG Name", "Type Line", "Storage Location"]:
                cell.alignment = left_align
                if col_name == "Final Fantasy Card Name":
                    cell.font = data_bold_font
            elif col_name in ["Card Image", "Card Set", "Card Number", "Print Variant",
                             "Available Variants", "Treatment / Frame", "Game Number", "Rarity",
                             "Color Identity", "Owned", "Condition"]:
                cell.alignment = center_align
            elif col_name in ["Non-Foil Qty", "Foil Qty", "Special Foil Qty", "Total Quantity Owned"]:
                cell.alignment = center_align
                cell.number_format = "0"
            elif col_name in ["Price (USD)", "Price (Foil)"]:
                cell.alignment = right_align
                cell.number_format = "$#,##0.00"
                
    max_row = total_rows + 1
    max_col = len(cols_to_write)
    max_col_letter = get_column_letter(max_col)
    
    ws_main.auto_filter.ref = f"A1:{max_col_letter}{max_row}"
    ws_main.freeze_panes = "A2"
    
    for col_idx, col_name in enumerate(cols_to_write, start=1):
        if col_name in val_col_ranges:
            col_letter = get_column_letter(col_idx)
            dv = DataValidation(
                type="list",
                formula1=val_col_ranges[col_name],
                allow_blank=True,
                showDropDown=False
            )
            ws_main.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}{max_row}")
            
    owned_col_idx = cols_to_write.index("Owned") + 1
    owned_col_letter = get_column_letter(owned_col_idx)
    
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    
    rule_yes = FormulaRule(formula=[f'${owned_col_letter}2="Yes"'], stopIfTrue=True, fill=green_fill)
    rule_no = FormulaRule(formula=[f'${owned_col_letter}2="No"'], stopIfTrue=True, fill=red_fill)
    
    ws_main.conditional_formatting.add(f"A2:{max_col_letter}{max_row}", rule_yes)
    ws_main.conditional_formatting.add(f"A2:{max_col_letter}{max_row}", rule_no)
    
    col_width_defaults = {
        "Final Fantasy Card Name": 30,
        "Original MTG Name": 24,
        "Card Image": 16,
        "Card Set": 12,
        "Card Number": 14,
        "Print Variant": 18,
        "Available Variants": 24,
        "Treatment / Frame": 18,
        "Game Number": 16,
        "Rarity": 14,
        "Type Line": 30,
        "Color Identity": 15,
        "Owned": 12,
        "Non-Foil Qty": 14,
        "Foil Qty": 14,
        "Special Foil Qty": 16,
        "Total Quantity Owned": 18,
        "Condition": 24,
        "Price (USD)": 15,
        "Price (Foil)": 15,
        "Storage Location": 22,
    }
    
    for col_idx, col_name in enumerate(cols_to_write, start=1):
        col_letter = get_column_letter(col_idx)
        ws_main.column_dimensions[col_letter].width = col_width_defaults.get(col_name, 18)
        
    print("Constructing 'Collection Dashboard' summary metrics and formulas...")
    
    navy_dark = "0F172A"
    slate_header = "1E293B"
    border_color = "CBD5E1"
    
    dash_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color)
    )
    
    ws_dash.merge_cells("A1:E1")
    title_cell = ws_dash["A1"]
    title_cell.value = "FINAL FANTASY MTG COLLECTION DASHBOARD"
    title_cell.font = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color=navy_dark, end_color=navy_dark, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[1].height = 40
    
    ws_dash.merge_cells("A2:E2")
    sub_cell = ws_dash["A2"]
    sub_cell.value = "Live Collection Statistics & Set Completion Tracking"
    sub_cell.font = Font(name="Segoe UI", size=10, italic=True, color="94A3B8")
    sub_cell.fill = PatternFill(start_color=slate_header, end_color=slate_header, fill_type="solid")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[2].height = 22
    
    ws_dash.row_dimensions[3].height = 14
    
    name_col = get_column_letter(cols_to_write.index("Final Fantasy Card Name") + 1)
    game_col = get_column_letter(cols_to_write.index("Game Number") + 1)
    owned_col = get_column_letter(cols_to_write.index("Owned") + 1)
    qty_col = get_column_letter(cols_to_write.index("Total Quantity Owned") + 1)
    price_col = get_column_letter(cols_to_write.index("Price (USD)") + 1)
    
    kpi_metrics = [
        ("Total Cards in Set", f"=COUNTA('FF MTG Collection'!{name_col}2:{name_col}{max_row})", "#,##0", "Total unique prints indexed from Scryfall"),
        ("Unique Cards Owned", f'=COUNTIF(\'FF MTG Collection\'!{owned_col}2:{owned_col}{max_row}, "Yes")', "#,##0", "Unique card variants in your collection"),
        ("Total Cards Quantity", f"=SUM('FF MTG Collection'!{qty_col}2:{qty_col}{max_row})", "#,##0", "Total count of all physical card copies across all variants"),
        ("Completion Rate (%)", "=B5/B4", "0.0%", "Overall unique set completion percentage"),
        ("Estimated Market Value", f"=SUMPRODUCT('FF MTG Collection'!{qty_col}2:{qty_col}{max_row}, 'FF MTG Collection'!{price_col}2:{price_col}{max_row})", "$#,##0.00", "Total estimated market value of owned cards"),
    ]
    
    for idx, (label, formula, num_format, desc) in enumerate(kpi_metrics, start=4):
        ws_dash.row_dimensions[idx].height = 26
        
        lbl_cell = ws_dash.cell(row=idx, column=1, value=label)
        lbl_cell.font = Font(name="Segoe UI", size=11, bold=True, color="1E293B")
        lbl_cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        lbl_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lbl_cell.border = dash_border
        
        val_cell = ws_dash.cell(row=idx, column=2, value=formula)
        val_cell.font = Font(name="Segoe UI", size=12, bold=True, color="0F172A")
        val_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        val_cell.alignment = Alignment(horizontal="right", vertical="center")
        val_cell.number_format = num_format
        val_cell.border = dash_border
        
        ws_dash.merge_cells(start_row=idx, start_column=3, end_row=idx, end_column=5)
        note_cell = ws_dash.cell(row=idx, column=3, value=desc)
        note_cell.font = Font(name="Segoe UI", size=9, italic=True, color="64748B")
        note_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for c in range(3, 6):
            ws_dash.cell(row=idx, column=c).border = dash_border
            ws_dash.cell(row=idx, column=c).fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
            
    ws_dash.row_dimensions[10].height = 18
    
    ws_dash.merge_cells("A11:E11")
    tbl_title = ws_dash["A11"]
    tbl_title.value = "COLLECTION PROGRESS BY FINAL FANTASY GAME"
    tbl_title.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    tbl_title.fill = PatternFill(start_color=slate_header, end_color=slate_header, fill_type="solid")
    tbl_title.alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[11].height = 28
    
    breakdown_headers = ["FF Game", "Total Cards", "Cards Owned", "Total Copies", "% Completed"]
    ws_dash.row_dimensions[12].height = 24
    for c_idx, h_text in enumerate(breakdown_headers, start=1):
        c = ws_dash.cell(row=12, column=c_idx, value=h_text)
        c.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = dash_border
        
    ff_games_list = VALIDATION_OPTIONS["Game Number"]
    start_row = 13
    
    for offset, game in enumerate(ff_games_list):
        current_r = start_row + offset
        ws_dash.row_dimensions[current_r].height = 22
        
        c_game = ws_dash.cell(row=current_r, column=1, value=game)
        c_game.font = Font(name="Segoe UI", size=10, bold=True, color="1E293B")
        c_game.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c_game.border = dash_border
        c_game.fill = PatternFill(start_color="F8FAFC" if offset % 2 == 0 else "FFFFFF",
                                 end_color="F8FAFC" if offset % 2 == 0 else "FFFFFF",
                                 fill_type="solid")
        
        c_total = ws_dash.cell(
            row=current_r, column=2,
            value=f"=COUNTIF('FF MTG Collection'!${game_col}$2:${game_col}${max_row}, A{current_r})"
        )
        c_total.font = Font(name="Segoe UI", size=10)
        c_total.alignment = Alignment(horizontal="right", vertical="center")
        c_total.number_format = "#,##0"
        c_total.border = dash_border
        
        c_owned = ws_dash.cell(
            row=current_r, column=3,
            value=f'=COUNTIFS(\'FF MTG Collection\'!${game_col}$2:${game_col}${max_row}, A{current_r}, \'FF MTG Collection\'!${owned_col}$2:${owned_col}${max_row}, "Yes")'
        )
        c_owned.font = Font(name="Segoe UI", size=10)
        c_owned.alignment = Alignment(horizontal="right", vertical="center")
        c_owned.number_format = "#,##0"
        c_owned.border = dash_border
        
        c_qty = ws_dash.cell(
            row=current_r, column=4,
            value=f"=SUMIFS('FF MTG Collection'!${qty_col}$2:${qty_col}${max_row}, 'FF MTG Collection'!${game_col}$2:${game_col}${max_row}, A{current_r})"
        )
        c_qty.font = Font(name="Segoe UI", size=10)
        c_qty.alignment = Alignment(horizontal="right", vertical="center")
        c_qty.number_format = "#,##0"
        c_qty.border = dash_border
        
        c_pct = ws_dash.cell(
            row=current_r, column=5,
            value=f"=IF(B{current_r}>0, C{current_r}/B{current_r}, 0)"
        )
        c_pct.font = Font(name="Segoe UI", size=10, bold=True, color="2563EB")
        c_pct.alignment = Alignment(horizontal="right", vertical="center")
        c_pct.number_format = "0.0%"
        c_pct.border = dash_border
        
    summary_r = start_row + len(ff_games_list)
    ws_dash.row_dimensions[summary_r].height = 26
    
    tot_label = ws_dash.cell(row=summary_r, column=1, value="Total / Overall")
    tot_label.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    tot_label.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    tot_label.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    tot_label.border = dash_border
    
    tot_b = ws_dash.cell(row=summary_r, column=2, value=f"=SUM(B{start_row}:B{summary_r - 1})")
    tot_b.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    tot_b.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    tot_b.alignment = Alignment(horizontal="right", vertical="center")
    tot_b.number_format = "#,##0"
    tot_b.border = dash_border
    
    tot_c = ws_dash.cell(row=summary_r, column=3, value=f"=SUM(C{start_row}:C{summary_r - 1})")
    tot_c.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    tot_c.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    tot_c.alignment = Alignment(horizontal="right", vertical="center")
    tot_c.number_format = "#,##0"
    tot_c.border = dash_border
    
    tot_d = ws_dash.cell(row=summary_r, column=4, value=f"=SUM(D{start_row}:D{summary_r - 1})")
    tot_d.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    tot_d.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    tot_d.alignment = Alignment(horizontal="right", vertical="center")
    tot_d.number_format = "#,##0"
    tot_d.border = dash_border
    
    tot_e = ws_dash.cell(row=summary_r, column=5, value=f"=IF(B{summary_r}>0, C{summary_r}/B{summary_r}, 0)")
    tot_e.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    tot_e.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    tot_e.alignment = Alignment(horizontal="right", vertical="center")
    tot_e.number_format = "0.0%"
    tot_e.border = dash_border
    
    ws_dash.column_dimensions["A"].width = 28
    ws_dash.column_dimensions["B"].width = 18
    ws_dash.column_dimensions["C"].width = 18
    ws_dash.column_dimensions["D"].width = 18
    ws_dash.column_dimensions["E"].width = 18
    
    print(f"\nSaving final workbook to '{output_path}'...")
    wb.save(output_path)
    print("Workbook successfully saved!")


# ---------------------------------------------------------------------------
# Main Execution Entry Point
# ---------------------------------------------------------------------------

def parse_args():
    parser = argparse.ArgumentParser(
        description="Build the Final Fantasy MTG collection tracker data files from Scryfall."
    )
    parser.add_argument(
        "--only-js", action="store_true",
        help="Only regenerate cards_data.js (what the website needs). Used by the monthly price refresh."
    )
    parser.add_argument(
        "--skip-js", action="store_true",
        help="Skip cards_data.js and build only the spreadsheet outputs."
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Write every output with a .new suffix so nothing existing is overwritten."
    )
    return parser.parse_args()


def main():
    args = parse_args()
    suffix = ".new" if args.dry_run else ""

    print("=" * 65)
    print("Magic: The Gathering - Final Fantasy Collection Tracker Builder")
    if args.dry_run:
        print("DRY RUN: writing *.new files; existing outputs are left alone.")
    print("=" * 65)

    try:
        raw_cards = fetch_scryfall_cards()
        if not raw_cards:
            print("No card data was returned from the API query.")
            sys.exit(1)

        written = []

        if not args.skip_js:
            records = build_card_records(raw_cards)
            js_path = OUTPUT_CARDS_JS + suffix
            export_cards_js(records, js_path)
            written.append(("Web card database", js_path))

            # Append today's prices. On a dry run this writes alongside rather
            # than into the real history, which must never be clobbered: unlike
            # every other output it cannot be rebuilt from Scryfall, because
            # Scryfall only serves today's price.
            history_path = OUTPUT_PRICE_HISTORY + suffix
            if args.dry_run and not os.path.exists(history_path):
                import shutil
                if os.path.exists(OUTPUT_PRICE_HISTORY):
                    shutil.copyfile(OUTPUT_PRICE_HISTORY, history_path)
            update_price_history(records, history_path)
            written.append(("Price history", history_path))

        if not args.only_js:
            df_cards = parse_card_data(raw_cards)
            xlsx_path = OUTPUT_XLSX + suffix
            csv_path = OUTPUT_CSV + suffix
            tsv_path = OUTPUT_TSV + suffix
            build_excel_tracker(df_cards, xlsx_path)
            export_csv_and_tsv(df_cards, csv_path, tsv_path)
            written.extend([
                ("Excel workbook", xlsx_path),
                ("CSV import file", csv_path),
                ("TSV import file", tsv_path),
            ])

        print("\n" + "=" * 65)
        print("SUCCESS: generated tracker files:")
        for index, (label, path) in enumerate(written, start=1):
            print(f"  {index}. {label}: {path}")
        print(f"Total cards: {len(raw_cards)}")
        print("=" * 65)

    except Exception as ex:
        print(f"\nAn error occurred during execution: {ex}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
